"""Seed / enrich Surgery rows from a ModMed appointments export (XLSX).

Match priority per row:
  1. By Patient MRN  → Surgery.chart_number (exact)
  2. By Email        → Surgery.email (case-insensitive)
  3. By First+Last+DOB
  4. By First+Last (and prompt the user if multiple match)

For matched rows: fill in any empty fields from ModMed but never overwrite
non-empty data unless the existing value is a Calendly TBD-* placeholder.
For unmatched rows: create a new Surgery with full ModMed data and status
'new' (not 'incomplete' — we have real MRN/DOB/phone here).

Dry-run by default. Pass --apply to commit.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date as _date, datetime, time as _time
from typing import Optional

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy import or_

from app.database import SessionLocal, init_db
from app.models.surgery import Surgery


APPT_TYPE_MAP = {
    "Office-Based Surgery": ("office",   "office"),
    "MedStar-Robot-Short":  ("medstar",  "robotic_180"),
    "MedStar-Robot-Long":   ("medstar",  "robotic_240"),
    "MedStar-Minor":        ("medstar",  "minor"),
    "CRMC-Major":           ("crmc",     "major"),
    "CRMC-Minor":           ("crmc",     "minor"),
}


def parse_time_str(s: str) -> Optional[_time]:
    if not s:
        return None
    s = s.strip()
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def is_tbd_chart(chart: Optional[str]) -> bool:
    return bool(chart and chart.startswith("TBD-"))


def title_word(word):
    """Conservative title case: only fixes fully UPPER or fully lower words."""
    if not word:
        return word
    alpha = re.sub(r"[^A-Za-z]", "", word)
    if not alpha or not (alpha.isupper() or alpha.islower()):
        return word
    return re.sub(r"[A-Za-z]+", lambda m: m.group(0).capitalize(), word)


def normalize_name(s):
    if not s:
        return s
    parts = re.split(r"(\s+|,)", s.strip())
    out = "".join(title_word(p) if p and not p.isspace() and p != "," else p for p in parts if p is not None)
    return re.sub(r"\s{2,}", " ", out).strip()


def find_match(db, row) -> tuple[str, list[Surgery]]:
    """Returns (match_kind, candidates)."""
    chart = (row.get("mrn") or "").strip()
    email = (row.get("email") or "").strip().lower()
    first = (row.get("first") or "").strip()
    last = (row.get("last") or "").strip()
    dob = row.get("dob")

    if chart:
        rows = db.query(Surgery).filter(Surgery.chart_number == chart).all()
        if len(rows) == 1: return "chart", rows
        if len(rows) > 1: return "chart_ambiguous", rows

    if email:
        rows = db.query(Surgery).filter(Surgery.email.ilike(email)).all()
        if len(rows) == 1: return "email", rows
        if len(rows) > 1:
            # Tiebreak with DOB if available
            if dob:
                m = [s for s in rows if s.dob == dob]
                if len(m) == 1: return "email_dob", m
            return "email_ambiguous", rows

    if first and last:
        candidates = (db.query(Surgery)
                        .filter(Surgery.patient_name.ilike(f"%{last}%"))
                        .all())
        candidates = [s for s in candidates
                      if first.lower() in (s.patient_name or "").lower()
                         and last.lower() in (s.patient_name or "").lower()]
        if dob:
            with_dob = [s for s in candidates if s.dob == dob]
            if len(with_dob) == 1: return "name_dob", with_dob
            if len(with_dob) > 1:  return "name_dob_ambiguous", with_dob
        if len(candidates) == 1: return "name", candidates
        if len(candidates) > 1:  return "name_ambiguous", candidates

    return "none", []


def coalesce(existing, incoming):
    """Return incoming when existing is None/empty/whitespace; else existing."""
    if existing is None or (isinstance(existing, str) and not existing.strip()):
        return incoming
    return existing


def enrich(s: Surgery, row, overwrite_chart_tbd=True) -> list[str]:
    """Fill empty fields from the ModMed row. Returns list of changed field names."""
    changes = []
    incoming = {
        "patient_name": normalize_name(f"{row.get('first', '')} {row.get('last', '')}".strip()),
        "first_name": normalize_name(row.get("first")),
        "last_name": normalize_name(row.get("last")),
        "dob": row.get("dob"),
        "email": (row.get("email") or "").lower() or None,
        "phone": row.get("phone"),
        "cell_phone": row.get("phone"),
        "address_street": row.get("addr1"),
        "address_city": normalize_name(row.get("city")),
        "address_state": (row.get("state") or "").upper() or None,
        "address_zip": (row.get("zip") or "").split("-")[0] or None,
    }
    # Schedule / facility
    appt_dt = row.get("appt_dt")
    appt_time = row.get("appt_time")
    fac, classification = APPT_TYPE_MAP.get(row.get("appt_type") or "", (None, None))

    for field, val in incoming.items():
        if val is None or (isinstance(val, str) and not val.strip()):
            continue
        cur = getattr(s, field, None)
        new = coalesce(cur, val)
        if new != cur:
            setattr(s, field, new)
            changes.append(field)

    # Chart number: overwrite TBD-* placeholders with real MRN
    if row.get("mrn") and overwrite_chart_tbd and is_tbd_chart(s.chart_number):
        s.chart_number = row["mrn"]
        changes.append("chart_number")

    # Schedule fields — only fill if empty (don't trample patient reschedules)
    if appt_dt and not s.scheduled_date:
        s.scheduled_date = appt_dt
        changes.append("scheduled_date")
    if appt_time and not s.scheduled_start_time:
        s.scheduled_start_time = appt_time
        changes.append("scheduled_start_time")
    if fac and not s.selected_facility:
        s.selected_facility = fac
        changes.append("selected_facility")
        el = list(s.eligible_facilities or [])
        if fac not in el:
            el.append(fac)
            s.eligible_facilities = el
            changes.append("eligible_facilities")
    if classification and not s.procedure_classification:
        s.procedure_classification = classification
        changes.append("procedure_classification")

    return changes


def create_from_modmed(row) -> Surgery:
    """Create a new Surgery row from a ModMed appointment."""
    fac, classification = APPT_TYPE_MAP.get(row.get("appt_type") or "", (None, None))
    first = normalize_name(row.get("first"))
    last = normalize_name(row.get("last"))
    return Surgery(
        chart_number=row.get("mrn"),
        patient_name=normalize_name(f"{first or ''} {last or ''}".strip()),
        first_name=first,
        last_name=last,
        dob=row.get("dob"),
        email=(row.get("email") or "").lower() or None,
        phone=row.get("phone"),
        cell_phone=row.get("phone"),
        address_street=row.get("addr1"),
        address_city=normalize_name(row.get("city")),
        address_state=(row.get("state") or "").upper() or None,
        address_zip=(row.get("zip") or "").split("-")[0] or None,
        scheduled_date=row.get("appt_dt"),
        scheduled_start_time=row.get("appt_time"),
        selected_facility=fac,
        eligible_facilities=[fac] if fac else [],
        procedure_classification=classification,
        # ModMed gives us enough data to spawn workflow milestones, so mark as new
        status="new",
        source="modmed",
        notes=f"Seeded from ModMed appointments export "
              f"(appt_type={row.get('appt_type')}, status={row.get('appt_status')}). "
              f"Procedure / insurance still need to be filled from chart.",
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    init_db()
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Build name → index map
        idx = {h: i for i, h in enumerate(header) if h}

        def cell(r, name):
            i = idx.get(name)
            if i is None: return None
            v = r[i]
            return v.strip() if isinstance(v, str) else v

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            dob = cell(r, "Patient DOB")
            appt_d = cell(r, "Appointment Date")
            rows.append({
                "mrn": str(cell(r, "Patient MRN") or "").strip() or None,
                "phone": cell(r, "Patient Mobile Phone"),
                "zip": cell(r, "Patient Zip Code"),
                "appt_time": parse_time_str(cell(r, "Appointment Time") or ""),
                "addr1": cell(r, "Patient Address Line 1"),
                "state": cell(r, "Patient State"),
                "email": cell(r, "Patient Email Address"),
                "appt_dt": appt_d.date() if isinstance(appt_d, datetime) else appt_d,
                "dob": dob.date() if isinstance(dob, datetime) else dob,
                "appt_status": cell(r, "Appointment Status"),
                "first": cell(r, "Patient First Name"),
                "appt_type": cell(r, "Appointment Type"),
                "last": cell(r, "Patient Last Name"),
                "city": cell(r, "Patient City"),
            })

        by_kind = {"chart": [], "email": [], "email_dob": [], "name_dob": [], "name": [],
                   "ambiguous": [], "none": []}
        update_plans = []  # (Surgery, row, changes_preview)
        creates = []
        ambiguous_rows = []

        for row in rows:
            kind, cands = find_match(db, row)
            if kind in ("chart", "email", "email_dob", "name_dob", "name"):
                # Preview changes without committing yet
                s = cands[0]
                snapshot = {
                    "patient_name": s.patient_name, "chart_number": s.chart_number,
                    "dob": s.dob, "email": s.email, "phone": s.phone,
                    "address_street": s.address_street, "scheduled_date": s.scheduled_date,
                    "selected_facility": s.selected_facility,
                }
                # Run enrich in dry-run (we'll roll back the in-memory mutation later)
                # For dry run we just compute what *would* change
                preview = []
                if row.get("mrn") and is_tbd_chart(s.chart_number):
                    preview.append(f"chart_number {s.chart_number} → {row['mrn']}")
                if row.get("dob") and not s.dob:
                    preview.append(f"dob → {row['dob']}")
                if row.get("phone") and not s.phone:
                    preview.append(f"phone → {row['phone']}")
                if row.get("addr1") and not s.address_street:
                    preview.append(f"address → {row['addr1']}")
                if row.get("appt_dt") and not s.scheduled_date:
                    preview.append(f"date → {row['appt_dt']}")
                if row.get("appt_time") and not s.scheduled_start_time:
                    preview.append(f"time → {row['appt_time']}")
                update_plans.append((s, row, kind, preview))
                by_kind[kind].append(row)
            elif kind in ("chart_ambiguous", "email_ambiguous", "name_ambiguous", "name_dob_ambiguous"):
                ambiguous_rows.append((row, kind, cands))
            else:
                creates.append(row)
                by_kind["none"].append(row)

        # ── REPORT ─────────────────────────────────────────────────
        print(f"\nModMed rows: {len(rows)}")
        print(f"  Matched by chart #:     {len(by_kind['chart'])}")
        print(f"  Matched by email:       {len(by_kind['email']) + len(by_kind['email_dob'])}")
        print(f"  Matched by name+DOB:    {len(by_kind['name_dob'])}")
        print(f"  Matched by name only:   {len(by_kind['name'])}")
        print(f"  AMBIGUOUS:              {len(ambiguous_rows)}")
        print(f"  Would CREATE:           {len(creates)}")

        if ambiguous_rows:
            print("\n── AMBIGUOUS ──")
            for row, kind, cands in ambiguous_rows:
                print(f"  {row['first']} {row['last']} chart={row['mrn']} email={row['email']} ({kind})")
                for c in cands:
                    print(f"    candidate chart={c.chart_number} name={c.patient_name!r} dob={c.dob}")

        tbd_upgrades = [(s, row) for s, row, k, p in update_plans
                        if any("chart_number TBD-" in x for x in p)]
        if tbd_upgrades:
            print(f"\n── Stub upgrades (TBD-* → real MRN) ──  {len(tbd_upgrades)}")
            for s, row in tbd_upgrades:
                print(f"  {s.patient_name}: {s.chart_number} → {row['mrn']}")

        if not args.apply:
            print(f"\nDRY RUN — no changes applied. Re-run with --apply.")
            return

        # ── APPLY ──────────────────────────────────────────────────
        n_enriched = 0
        for s, row, kind, preview in update_plans:
            changes = enrich(s, row)
            if changes:
                n_enriched += 1

        for row in creates:
            db.add(create_from_modmed(row))

        db.commit()
        print(f"\n✓ Enriched {n_enriched} existing surgeries.")
        print(f"✓ Created {len(creates)} new surgeries from ModMed.")
        print(f"⚠ Skipped {len(ambiguous_rows)} ambiguous rows — review manually.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
