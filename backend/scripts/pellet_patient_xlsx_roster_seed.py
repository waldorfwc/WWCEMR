"""Seed / upsert pellet patient roster from a ModMed Qlik export (.xlsx).

The export typically has one row per (patient × payer plan), so the same
MRN can appear multiple times — we dedupe by MRN and take the FIRST
payer encountered as the primary_insurance.

Columns expected (case-sensitive header match in row 1):
  • Patient MRN              → chart_number
  • Patient First Name       \\
  • Patient Last Name        /  → patient_name = "First Last"
  • Patient Link             → modmed_link (Qlik redirect URL)
  • Patient DOB              → patient_dob
  • Patient Email Address    → patient_email
  • Payer Name               → primary_insurance (first occurrence per MRN)

Behavior:
  • Match an existing PelletPatient by chart_number → UPDATE (filling
    fields, never clobbering with empty values).
  • No match → INSERT a new pellet patient with patient_type='established'
    (since these are real ModMed patients pulled from the BHRT roster).
  • Idempotent: re-running only changes rows that drift.

Usage:
  ./venv/bin/python scripts/pellet_patient_xlsx_roster_seed.py PATH [--dry-run]
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from typing import Optional

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.models.pellet import PelletPatient  # noqa: E402


REQUIRED_HEADERS = {
    "Patient MRN":            "chart_number",
    "Patient First Name":     "first_name",
    "Patient Last Name":      "last_name",
    "Patient Link":           "modmed_link",
    "Patient DOB":            "patient_dob",
    "Patient Email Address":  "patient_email",
    "Payer Name":             "primary_insurance",
}


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("spreadsheet is empty")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    missing = [h for h in REQUIRED_HEADERS if h not in header]
    if missing:
        raise RuntimeError(f"missing required headers: {missing}\nfound: {header}")
    col_idx = {h: header.index(h) for h in REQUIRED_HEADERS}
    out = []
    for raw in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        rec = {key: _clean(raw[col_idx[h]]) for h, key in REQUIRED_HEADERS.items()}
        # Convert DOB
        rec["patient_dob"] = _to_date(raw[col_idx["Patient DOB"]])
        out.append(rec)
    return out


def dedupe_by_mrn(rows: list[dict]) -> list[dict]:
    """Collapse duplicate MRN rows; keep the first encountered payer."""
    seen = {}
    for r in rows:
        mrn = r.get("chart_number")
        if not mrn:
            continue
        if mrn in seen:
            # If the existing entry lacks any field, fill it from the dup
            cur = seen[mrn]
            for k, v in r.items():
                if v and not cur.get(k):
                    cur[k] = v
            continue
        seen[mrn] = dict(r)
    return list(seen.values())


JUNK_INSURANCE = {"none", "-", "patient", "self-pay", "self pay", "n/a", "na"}


def _is_junk_insurance(v: Optional[str]) -> bool:
    return not v or v.strip().lower() in JUNK_INSURANCE


def _name_covers(existing: Optional[str], first: Optional[str],
                  last: Optional[str]) -> bool:
    """True when the existing patient_name already contains both the
    spreadsheet's first and last names (case-insensitive). Avoids the
    'Atkinson, Venus D.' → 'Venus Atkinson' reformat that drops the
    middle initial."""
    if not existing or not first or not last:
        return False
    e = existing.lower()
    return first.lower() in e and last.lower() in e


def apply(rows: list[dict], *, dry_run: bool, actor: str) -> dict:
    db = SessionLocal()
    summary = {"created": 0, "updated": 0, "unchanged": 0,
               "name_preserved": 0, "insurance_skipped": 0, "changes": []}
    try:
        existing_by_chart = {p.chart_number: p for p in db.query(PelletPatient).all()}
        for r in rows:
            chart = r["chart_number"]
            if not chart:
                continue
            first = r.get("first_name")
            last  = r.get("last_name")
            spread_name = " ".join([first or "", last or ""]).strip() or None

            row = existing_by_chart.get(chart)
            if row is None:
                if not spread_name:
                    print(f"  [skip] MRN {chart}: no name in spreadsheet")
                    continue
                summary["created"] += 1
                ins = r.get("primary_insurance")
                if _is_junk_insurance(ins):
                    ins = None
                summary["changes"].append(f"CREATE {chart}: {spread_name}")
                if not dry_run:
                    db.add(PelletPatient(
                        chart_number=chart,
                        patient_name=spread_name,
                        patient_dob=r.get("patient_dob"),
                        patient_email=r.get("patient_email"),
                        primary_insurance=ins,
                        modmed_link=r.get("modmed_link"),
                        patient_type="established",
                        created_by=actor,
                    ))
                continue

            # UPDATE existing — fill fields that meaningfully improve the row.
            changes = {}

            # patient_name: only overwrite when the existing name doesn't
            # already cover both first and last (preserves middle initials).
            if spread_name and not _name_covers(row.patient_name, first, last):
                if row.patient_name != spread_name:
                    changes["patient_name"] = (row.patient_name, spread_name)
            elif spread_name and row.patient_name and row.patient_name != spread_name:
                summary["name_preserved"] += 1

            # patient_dob / patient_email — fill if differs and non-empty
            if r.get("patient_dob") and r["patient_dob"] != row.patient_dob:
                changes["patient_dob"] = (row.patient_dob, r["patient_dob"])
            if r.get("patient_email") and r["patient_email"] != row.patient_email:
                changes["patient_email"] = (row.patient_email, r["patient_email"])

            # primary_insurance — skip junk values from the spreadsheet
            ins = r.get("primary_insurance")
            if ins:
                if _is_junk_insurance(ins):
                    summary["insurance_skipped"] += 1
                elif ins != row.primary_insurance:
                    changes["primary_insurance"] = (row.primary_insurance, ins)

            # modmed_link — always fill if differs (this is the main point)
            if r.get("modmed_link") and r["modmed_link"] != row.modmed_link:
                changes["modmed_link"] = (row.modmed_link, r["modmed_link"])

            if not changes:
                summary["unchanged"] += 1
                continue
            summary["updated"] += 1
            summary["changes"].append(
                f"UPDATE {chart} ({row.patient_name}): "
                + ", ".join(f"{k}={a!r}→{b!r}" for k, (a, b) in changes.items())
            )
            if not dry_run:
                for k, (_, b) in changes.items():
                    setattr(row, k, b)
        if not dry_run:
            db.commit()
    finally:
        db.close()
    return summary


def main():
    ap = argparse.ArgumentParser(description="Upsert pellet patients from xlsx roster")
    ap.add_argument("path", help="path to the .xlsx export")
    ap.add_argument("--dry-run", action="store_true",
                    help="show what would change without writing")
    ap.add_argument("--actor", default="system:xlsx-roster",
                    help="created_by attribution string")
    args = ap.parse_args()

    if not os.path.exists(args.path):
        print(f"file not found: {args.path}", file=sys.stderr)
        sys.exit(2)

    raw = load_rows(args.path)
    unique = dedupe_by_mrn(raw)
    print(f"Loaded {len(raw)} rows from {args.path}; {len(unique)} unique MRNs")
    print()
    summary = apply(unique, dry_run=args.dry_run, actor=args.actor)
    print()
    print(f"  Created:           {summary['created']}")
    print(f"  Updated:           {summary['updated']}")
    print(f"  Unchanged:         {summary['unchanged']}")
    print(f"  Name preserved:    {summary['name_preserved']} "
          f"(spreadsheet ≠ existing but existing already covers first+last)")
    print(f"  Insurance skipped: {summary['insurance_skipped']} "
          f"(spreadsheet value was 'None'/'Patient'/'-' etc.)")
    print()
    for line in summary["changes"]:
        print(f"  • {line}")
    if args.dry_run:
        print("\n(dry-run — no rows committed)")


if __name__ == "__main__":
    main()
