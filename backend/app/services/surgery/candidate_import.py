"""Parse a ModMed-style patient roster Excel and create Surgery rows in
'incomplete' status — so coordinators can triage who needs benefits
checks, prior auth outreach, etc.

Expected columns (case- and whitespace-insensitive):
  Patient MRN
  Patient First Name
  Patient Last Name
  Patient DOB
  Patient Mobile Phone
  Patient Email Address
  Patient Address Line 1 / 2
  Patient City / State / Zip Code
  Payer
  Payer Plan Name
  Payer Policy Number
  Secondary Payer
  Tertiary Payer
  Primary Care Provider
  Appointment Count
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime, time as _time
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.surgery import Surgery

log = logging.getLogger(__name__)


# ─── header normalization ──────────────────────────────────────────

# Map normalized header text → canonical field key. Whitespace / case
# differences are stripped before lookup so minor variations don't break
# imports.
_HEADER_MAP = {
    "patientmrn":            "mrn",
    "patientfirstname":      "first",
    "patientlastname":       "last",
    "patientdob":            "dob",
    "patientmobilephone":    "phone",
    "patientemailaddress":   "email",
    "patientaddressline1":   "addr1",
    "patientaddressline2":   "addr2",
    "patientcity":           "city",
    "patientstate":          "state",
    "patientzipcode":        "zip",
    "payer":                 "payer",
    "payerplanname":         "payer_plan",
    "payerpolicynumber":     "payer_policy",
    "secondarypayer":        "payer_secondary",
    "tertiarypayer":         "payer_tertiary",
    "primarycareprovider":   "pcp",
    "appointmenttype":       "appt_type",
    "appointmentdate":       "appt_date",
    "appointmenttime":       "appt_time",
    "appointmentcount":      "appt_count",
}


# Map ModMed "Appointment Type" → internal (facility, procedure_kind, duration).
# The user has asked us to keep the appointment type as the procedure name on
# each surgery row (these are ModMed's generic labels and will be renamed by
# hand later), so this table only resolves the booking attributes.
APPT_TYPE_MAP: dict[str, tuple[str, str, int]] = {
    "medstar-robot-short":   ("medstar", "robotic_180", 180),
    "medstar-robot-long":    ("medstar", "robotic_240", 240),
    "medstar-minor":         ("medstar", "minor",        60),
    "crmc-major":            ("crmc",    "major",       120),
    "office-based surgery":  ("office",  "office",       30),
}


def _resolve_appt_type(label: Optional[str]) -> Optional[tuple[str, str, int]]:
    """Case-insensitive lookup. Returns None if the label isn't a known type."""
    if not label:
        return None
    return APPT_TYPE_MAP.get(label.strip().lower())


def _norm_header(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _clean_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in {"-", "—", "None", "none", "N/A", "NULL"}:
        return None
    return s


def _digits_only(v) -> Optional[str]:
    s = _clean_str(v)
    if not s:
        return None
    digits = re.sub(r"\D", "", s)
    return digits or None


def _as_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _as_time(v) -> Optional[_time]:
    """Parse strings like '1:30 PM' / '13:30' / '7:30 AM' from the ModMed
    roster. Returns None on anything unparseable."""
    if v is None or v == "":
        return None
    if isinstance(v, _time):
        return v
    if isinstance(v, datetime):
        return v.time()
    s = str(v).strip()
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


# ─── parsing ────────────────────────────────────────────────────────

def parse_excel(content: bytes) -> list[dict]:
    """Return a list of normalized {field: value} dicts. Empty/blank rows
    are dropped. Skips rows with no MRN."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(values_only=True), None)
    if not header_row:
        return []
    col_map: dict[int, str] = {}
    for i, cell in enumerate(header_row):
        key = _HEADER_MAP.get(_norm_header(cell))
        if key:
            col_map[i] = key

    if "mrn" not in col_map.values():
        raise ValueError("Could not find a 'Patient MRN' column.")

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec: dict[str, Any] = {}
        for i, val in enumerate(row):
            key = col_map.get(i)
            if not key:
                continue
            rec[key] = val
        mrn = _clean_str(rec.get("mrn"))
        if not mrn:
            continue
        out.append(rec)
    return out


def _format_patient_name(rec: dict) -> str:
    first = (_clean_str(rec.get("first")) or "").title()
    last  = (_clean_str(rec.get("last"))  or "").title()
    if last and first:
        return f"{last}, {first}"
    return last or first or "Unknown"


def _street(rec: dict) -> Optional[str]:
    a1 = _clean_str(rec.get("addr1"))
    a2 = _clean_str(rec.get("addr2"))
    if a1 and a2 and a2 != a1:
        return f"{a1}, {a2}"
    return a1 or a2


def _insurance(rec: dict) -> Optional[str]:
    """Prefer the plan name when distinct from the carrier; else use the
    carrier name. ModMed often splits 'BCBS National PPO Plan' (carrier)
    and 'BCBS PPO' (plan) — we keep the more descriptive of the two."""
    carrier = _clean_str(rec.get("payer"))
    plan    = _clean_str(rec.get("payer_plan"))
    if carrier and plan and plan.lower() not in carrier.lower():
        return f"{carrier} — {plan}"
    return carrier or plan


def _rec_to_surgery_kwargs(rec: dict) -> dict:
    appt_label = _clean_str(rec.get("appt_type"))
    appt_info  = _resolve_appt_type(appt_label)
    out: dict[str, Any] = {
        "chart_number":        _clean_str(rec["mrn"]),
        "first_name":          (_clean_str(rec.get("first")) or "").title() or None,
        "last_name":           (_clean_str(rec.get("last"))  or "").title() or None,
        "patient_name":        _format_patient_name(rec),
        "dob":                 _as_date(rec.get("dob")),
        "phone":               _digits_only(rec.get("phone")),
        "cell_phone":          _digits_only(rec.get("phone")),
        "email":               (_clean_str(rec.get("email")) or "").lower() or None,
        "address_street":      _street(rec),
        "address_city":        _clean_str(rec.get("city")),
        "address_state":       _clean_str(rec.get("state")),
        "address_zip":         _clean_str(rec.get("zip")),
        "primary_insurance":   _insurance(rec),
        "primary_member_id":   _clean_str(rec.get("payer_policy")),
        "secondary_insurance": _clean_str(rec.get("payer_secondary")),
        "status":              "incomplete",
        "sub_flag":            "candidate_imported",
    }
    # Keep the ModMed Appointment Type as the procedure name — the
    # coordinator will rename it to the real procedure later from the
    # SurgeryDetail screen. Also stamp the booking intent so the import
    # can attempt to book the slot if auto_schedule is on, AND so the
    # date-picker modal works after the fact for rows that didn't
    # auto-book (every booking flow gates on procedure_classification +
    # eligible_facilities).
    if appt_label:
        out["procedures"] = [{"name": appt_label}]
    if appt_info:
        facility, procedure_kind, duration = appt_info
        out["selected_facility"]         = facility
        out["eligible_facilities"]       = [facility]
        out["procedure_classification"]  = procedure_kind
        out["duration_minutes"]          = duration
        # Surface the robotic flag so dashboard / filters / capacity
        # rules treat the row the same as a UI-entered robotic surgery.
        out["is_robotic"] = procedure_kind in ("robotic_180", "robotic_240")
    return out


# ─── import ─────────────────────────────────────────────────────────

def import_rows(db: Session, rows: list[dict], *,
                dry_run: bool, by_email: str,
                auto_schedule: bool = False) -> dict:
    """Create Surgery rows in 'incomplete' status. Skips a row when an
    open (non-cancelled, non-completed) Surgery already exists for the
    same chart_number — coordinators can clean up duplicates by hand.

    When auto_schedule is True, every imported row that also carries an
    appointment_type + date + time is booked onto the matching BlockDay
    via book_slot — no patient email/SMS/calendar sync (the coordinator
    confirmed the dates externally).
    """
    from app.models.surgery import BlockDay, SurgeryNote

    created: list[dict] = []
    skipped: list[dict] = []
    errors:  list[dict] = []
    scheduled_summary: list[dict] = []
    schedule_errors: list[dict] = []

    for rec in rows:
        try:
            kwargs = _rec_to_surgery_kwargs(rec)
            chart = kwargs["chart_number"]
            appt_label = _clean_str(rec.get("appt_type"))
            appt_date  = _as_date(rec.get("appt_date"))
            appt_time  = _as_time(rec.get("appt_time"))
            appt_info  = _resolve_appt_type(appt_label)

            # Dedupe — any active surgery for this chart blocks creation
            existing = (db.query(Surgery)
                          .filter(Surgery.chart_number == chart,
                                  Surgery.status.notin_(["cancelled", "completed"]))
                          .first())
            if existing:
                skipped.append({
                    "chart_number": chart,
                    "patient_name": kwargs["patient_name"],
                    "reason": f"already has an active surgery ({existing.status})",
                })
                continue

            if dry_run:
                created.append({
                    "chart_number":      chart,
                    "patient_name":      kwargs["patient_name"],
                    "dob":               str(kwargs["dob"]) if kwargs["dob"] else None,
                    "primary_insurance": kwargs["primary_insurance"],
                    "appointment_type":  appt_label,
                    "appointment_date":  str(appt_date) if appt_date else None,
                    "appointment_time":  str(appt_time)[:5] if appt_time else None,
                    "would_auto_schedule": bool(
                        auto_schedule and appt_info and appt_date and appt_time),
                })
                continue

            s = Surgery(**kwargs)
            db.add(s)
            db.flush()
            created.append({
                "id":                str(s.id),
                "chart_number":      chart,
                "patient_name":      kwargs["patient_name"],
                "dob":               str(kwargs["dob"]) if kwargs["dob"] else None,
                "primary_insurance": kwargs["primary_insurance"],
            })

            # Optional auto-book against an existing BlockDay. Silent —
            # no confirmation email, no SMS, no calendar sync.
            if not (auto_schedule and appt_info and appt_date and appt_time):
                continue
            facility, procedure_kind, default_duration = appt_info
            bd = (db.query(BlockDay)
                    .filter(BlockDay.block_date == appt_date,
                            BlockDay.facility == facility)
                    .first())
            if not bd:
                schedule_errors.append({
                    "chart_number": chart,
                    "patient_name": kwargs["patient_name"],
                    "reason": (f"No BlockDay for {appt_date} at {facility} "
                                f"— set up the block schedule first."),
                })
                continue
            try:
                from app.services.surgery.block_schedule import (
                    book_slot, CapacityViolation,
                )
                slot = book_slot(
                    db, block_day_id=str(bd.id), surgery_id=str(s.id),
                    start_time=appt_time,
                    duration_minutes=default_duration,
                    procedure_kind=procedure_kind,
                )
            except CapacityViolation as exc:
                schedule_errors.append({
                    "chart_number": chart,
                    "patient_name": kwargs["patient_name"],
                    "reason":       str(exc),
                })
                continue
            # book_slot only auto-confirms when prior status was 'new' or
            # 'in_progress' — but the importer created this row as
            # 'incomplete', so the status stayed 'incomplete' and the
            # surgery dropped out of the calendar/list filters
            # (`Surgery.status.in_(("new","in_progress","confirmed"))`).
            # Explicitly bump to 'confirmed' here since we just booked a
            # real slot for this row.
            if s.status == "incomplete":
                s.status = "confirmed"
            db.add(SurgeryNote(
                surgery_id=s.id,
                created_by=by_email,
                content=(f"Bulk-import silent-scheduled {appt_date} "
                          f"{appt_time.strftime('%H:%M')} "
                          f"({default_duration} min) at {facility}. "
                          f"Procedure label from ModMed: {appt_label}."),
            ))
            scheduled_summary.append({
                "chart_number": chart,
                "patient_name": kwargs["patient_name"],
                "block_date":   str(appt_date),
                "start_time":   appt_time.strftime("%H:%M"),
                "duration":     default_duration,
                "facility":     facility,
                "procedure":    appt_label,
            })
        except Exception as exc:
            log.exception("candidate import row failed")
            errors.append({
                "chart_number": _clean_str(rec.get("mrn")) or "(missing)",
                "error": str(exc)[:200],
            })

    if not dry_run:
        db.commit()

    return {
        "total":   len(rows),
        "created": len(created),
        "skipped": len(skipped),
        "errors":  len(errors),
        "scheduled":       len(scheduled_summary),
        "schedule_errors": len(schedule_errors),
        "created_rows":    created[:200],
        "skipped_rows":    skipped[:200],
        "error_rows":      errors[:200],
        "scheduled_rows":  scheduled_summary[:200],
        "schedule_error_rows": schedule_errors[:200],
        "dry_run":         dry_run,
        "auto_schedule":   auto_schedule,
        "by":              by_email,
    }
