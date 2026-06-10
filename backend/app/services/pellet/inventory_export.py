"""Pellet inventory export: Excel (xlsx) + PDF.

Both render the same rolled-up view shown on the Lots in inventory card:
one row per lot, columns for total + per-location balances. Grouped by
dose-type label.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

LOC_LABEL = {
    "white_plains": "White Plains",
    "brandywine":   "Brandywine",
    "arlington":    "Arlington",
}
LOC_ORDER = ["white_plains", "brandywine", "arlington"]


def _group_rows(lots: list[dict]) -> list[dict]:
    """Group lot dicts by dose_type_label, sorted estradiol→testosterone,
    then alphanumerically. Returns one dict per group with `lots` inside."""
    by: dict[str, dict] = {}
    for l in lots:
        key = l.get("dose_type_label") or "(unknown)"
        if key not in by:
            by[key] = {
                "dose_type_label": key,
                "hormone": l.get("hormone") or "",
                "is_controlled": bool(l.get("is_controlled")),
                "lots": [],
                "total_on_hand": 0,
                "balances": {loc: 0 for loc in LOC_ORDER},
            }
        g = by[key]
        bal = l.get("balances") or {}
        total = sum(bal.values()) if bal else (l.get("total_on_hand") or 0)
        g["lots"].append(l)
        g["total_on_hand"] += total
        for loc in LOC_ORDER:
            g["balances"][loc] += bal.get(loc, 0)

    out = list(by.values())
    out.sort(key=lambda g: (
        0 if g["hormone"] == "estradiol" else 1 if g["hormone"] == "testosterone" else 2,
        g["dose_type_label"],
    ))
    return out


# ─── XLSX ────────────────────────────────────────────────────────────

def build_xlsx(lots: list[dict], *,
                filters_meta: Optional[dict] = None,
                generated_by: str = "system") -> bytes:
    """Render the inventory as a multi-section Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pellet Inventory"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="7B2D5E")
    header_font = Font(bold=True, color="FFFFFF")
    subhead_fill = PatternFill("solid", fgColor="F3E8EF")
    thin = Side(border_style="thin", color="D6C9D2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Title block
    ws.cell(row=1, column=1, value="WWC Pellet Inventory").font = Font(bold=True, size=14, color="7B2D5E")
    meta_bits = [f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"]
    if generated_by and generated_by != "system":
        meta_bits.append(f"by {generated_by}")
    if filters_meta:
        for k, v in filters_meta.items():
            if v:
                meta_bits.append(f"{k}={v}")
    ws.cell(row=2, column=1, value=" · ".join(meta_bits)).font = Font(italic=True, color="666666")

    # Header
    headers = ["Dose Type", "Lot #", "Expires", "White Plains", "Brandywine", "Arlington", "Total"]
    hdr_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    today_iso = date.today().isoformat()
    r = hdr_row + 1
    groups = _group_rows(lots)
    grand = {loc: 0 for loc in LOC_ORDER}
    grand_total = 0

    for g in groups:
        # Group subtotal row
        ws.cell(row=r, column=1, value=g["dose_type_label"]).font = bold
        ws.cell(row=r, column=2, value="(subtotal)").font = Font(italic=True, color="888888")
        ws.cell(row=r, column=3, value="")
        for i, loc in enumerate(LOC_ORDER, start=4):
            cell = ws.cell(row=r, column=i, value=g["balances"][loc])
            cell.font = bold
            cell.alignment = center
        total_cell = ws.cell(row=r, column=7, value=g["total_on_hand"])
        total_cell.font = bold
        total_cell.alignment = center
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = subhead_fill
            ws.cell(row=r, column=c).border = border
        grand_total += g["total_on_hand"]
        for loc in LOC_ORDER:
            grand[loc] += g["balances"][loc]
        r += 1

        # Detail rows
        for l in sorted(g["lots"], key=lambda x: (x.get("expiration_date") or "")):
            bal = l.get("balances") or {}
            ws.cell(row=r, column=1, value="")
            ws.cell(row=r, column=2, value=l.get("qualgen_lot_number") or "")
            exp = l.get("expiration_date") or ""
            exp_cell = ws.cell(row=r, column=3, value=exp)
            if exp and exp < today_iso:
                exp_cell.font = Font(color="C0392B", italic=True)
            for i, loc in enumerate(LOC_ORDER, start=4):
                v = bal.get(loc, 0)
                cell = ws.cell(row=r, column=i, value=v)
                cell.alignment = center
                if v == 0:
                    cell.font = Font(color="CCCCCC")
            ws.cell(row=r, column=7,
                    value=sum(bal.values())).alignment = center
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = border
            r += 1

    # Grand total
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = Font(bold=True, color="7B2D5E")
    for i, loc in enumerate(LOC_ORDER, start=4):
        cell = ws.cell(row=r, column=i, value=grand[loc])
        cell.font = Font(bold=True, color="7B2D5E")
        cell.alignment = center
    cell = ws.cell(row=r, column=7, value=grand_total)
    cell.font = Font(bold=True, color="7B2D5E")
    cell.alignment = center
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="EFE3EA")
        ws.cell(row=r, column=c).border = border

    # Column widths
    widths = [28, 18, 12, 14, 14, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF ─────────────────────────────────────────────────────────────

def build_pdf(lots: list[dict], *,
               filters_meta: Optional[dict] = None,
               generated_by: str = "system") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                              leftMargin=0.4 * inch, rightMargin=0.4 * inch,
                              topMargin=0.4 * inch, bottomMargin=0.4 * inch,
                              title="WWC Pellet Inventory")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                  textColor=colors.HexColor("#7B2D5E"),
                                  fontSize=16, spaceAfter=4)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"],
                                  fontSize=8, textColor=colors.gray)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"],
                                  fontSize=8)

    story = []
    story.append(Paragraph("WWC Pellet Inventory", title_style))

    meta_bits = [datetime.now().strftime("Generated %Y-%m-%d %H:%M")]
    if generated_by and generated_by != "system":
        meta_bits.append(f"by {generated_by}")
    if filters_meta:
        for k, v in filters_meta.items():
            if v:
                meta_bits.append(f"{k}={v}")
    story.append(Paragraph(" · ".join(meta_bits), meta_style))
    story.append(Spacer(1, 0.1 * inch))

    headers = ["Dose Type", "Lot #", "Expires",
                "White Plains", "Brandywine", "Arlington", "Total"]
    data = [headers]

    today_iso = date.today().isoformat()
    groups = _group_rows(lots)
    grand = {loc: 0 for loc in LOC_ORDER}
    grand_total = 0
    section_rows = []  # row indices that are group-subtotals (for styling)
    expired_rows = []  # row indices with an expired lot

    for g in groups:
        # Subtotal row
        data.append([
            g["dose_type_label"], "(subtotal)", "",
            str(g["balances"]["white_plains"]),
            str(g["balances"]["brandywine"]),
            str(g["balances"]["arlington"]),
            str(g["total_on_hand"]),
        ])
        section_rows.append(len(data) - 1)
        grand_total += g["total_on_hand"]
        for loc in LOC_ORDER:
            grand[loc] += g["balances"][loc]

        # Detail rows
        for l in sorted(g["lots"], key=lambda x: (x.get("expiration_date") or "")):
            bal = l.get("balances") or {}
            exp = l.get("expiration_date") or ""
            row = [
                "", l.get("qualgen_lot_number") or "", exp,
                str(bal.get("white_plains", 0)),
                str(bal.get("brandywine", 0)),
                str(bal.get("arlington", 0)),
                str(sum(bal.values())),
            ]
            data.append(row)
            if exp and exp < today_iso:
                expired_rows.append(len(data) - 1)

    # Grand total
    data.append([
        "GRAND TOTAL", "", "",
        str(grand["white_plains"]), str(grand["brandywine"]),
        str(grand["arlington"]), str(grand_total),
    ])
    grand_row = len(data) - 1

    col_widths = [2.2 * inch, 1.6 * inch, 0.9 * inch,
                   1.3 * inch, 1.1 * inch, 1.1 * inch, 0.8 * inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7B2D5E")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ALIGN",      (3, 1), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#D6C9D2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
            [colors.white, colors.HexColor("#FAF7F9")]),
    ])
    # Subtotal rows
    for i in section_rows:
        style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F3E8EF"))
        style.add("FONTNAME",   (0, i), (-1, i), "Helvetica-Bold")
    # Expired rows
    for i in expired_rows:
        style.add("TEXTCOLOR", (2, i), (2, i), colors.HexColor("#C0392B"))
    # Grand total
    style.add("BACKGROUND", (0, grand_row), (-1, grand_row), colors.HexColor("#EFE3EA"))
    style.add("FONTNAME",   (0, grand_row), (-1, grand_row), "Helvetica-Bold")
    style.add("TEXTCOLOR",  (0, grand_row), (-1, grand_row), colors.HexColor("#7B2D5E"))

    table.setStyle(style)
    story.append(table)

    doc.build(story)
    return buf.getvalue()
